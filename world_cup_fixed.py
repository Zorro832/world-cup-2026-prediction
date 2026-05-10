from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 定义样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
normal_font = Font(size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. 使用说明
ws_guide = wb.active
ws_guide.title = "使用说明"
ws_guide.column_dimensions['A'].width = 80

guide_text = [
    "2026年足球世界杯竞猜活动 - 使用说明",
    "",
    "【积分规则】",
    "• 预测准确比分：得 4 分",
    "• 只猜中胜负平关系：得 1 分",
    "• 胜负关系和比分都没猜中：得 0 分",
    "",
    "【工作表说明】",
    "1. 比赛赛程 - 管理员填写所有104场比赛信息和实际比分",
    "2. 参与者填写 - 每位参与者填写预测比分",
    "3. 积分统计 - 自动计算每位参与者每场比赛的积分",
    "4. 总排名 - 自动生成最终排名",
    "",
    "【使用步骤】",
    "1. 在"比赛赛程"表中填写所有比赛的实际比分",
    "2. 将文件分享到微信群，参与者在"参与者填写"表中填写预测",
    "3. 比赛结束后填写实际比分，积分和排名会自动更新",
    "",
    "注意：本表格预设104场比赛（小组赛72场+淘汰赛32场）"
]

for row, text in enumerate(guide_text, 1):
    cell = ws_guide.cell(row=row, column=1, value=text)
    if row == 1:
        cell.font = Font(bold=True, size=16, color='366092')
        ws_guide.row_dimensions[row].height = 35
    elif text.startswith("【"):
        cell.font = Font(bold=True, size=12, color='366092')
        ws_guide.row_dimensions[row].height = 25
    else:
        cell.font = Font(size=10)
        ws_guide.row_dimensions[row].height = 20
    cell.alignment = Alignment(vertical='center', wrap_text=True)

# 2. 比赛赛程（104场比赛）
ws_schedule = wb.create_sheet("比赛赛程")
headers = ["比赛编号", "比赛阶段", "日期", "队伍A", "队伍B", "实际比分A", "实际比分B"]
ws_schedule.row_dimensions[1].height = 30

for col, header in enumerate(headers, 1):
    cell = ws_schedule.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws_schedule.column_dimensions['A'].width = 12
ws_schedule.column_dimensions['B'].width = 25
ws_schedule.column_dimensions['C'].width = 15
ws_schedule.column_dimensions['D'].width = 20
ws_schedule.column_dimensions['E'].width = 20
ws_schedule.column_dimensions['F'].width = 12
ws_schedule.column_dimensions['G'].width = 12

# 创建104场比赛
matches = []

# 小组赛：12组×6场 = 72场
groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for group in groups:
    for match in range(1, 7):
        match_id = len(matches) + 1
        matches.append([
            match_id,
            f"小组赛{group}组第{match}轮",
            f"2026-06-{10 + (match_id-1)//8:02d}",
            f"{group}队1",
            f"{group}队2",
            "",
            ""
        ])

# 淘汰赛：32场
knockout = [
    (73, "Round of 32 - 第1场", "2026-06-28", "待定", "待定", "", ""),
    (74, "Round of 32 - 第2场", "2026-06-28", "待定", "待定", "", ""),
    (75, "Round of 32 - 第3场", "2026-06-28", "待定", "待定", "", ""),
    (76, "Round of 32 - 第4场", "2026-06-28", "待定", "待定", "", ""),
    (77, "Round of 32 - 第5场", "2026-06-29", "待定", "待定", "", ""),
    (78, "Round of 32 - 第6场", "2026-06-29", "待定", "待定", "", ""),
    (79, "Round of 32 - 第7场", "2026-06-29", "待定", "待定", "", ""),
    (80, "Round of 32 - 第8场", "2026-06-29", "待定", "待定", "", ""),
    (81, "Round of 32 - 第9场", "2026-06-30", "待定", "待定", "", ""),
    (82, "Round of 32 - 第10场", "2026-06-30", "待定", "待定", "", ""),
    (83, "Round of 32 - 第11场", "2026-06-30", "待定", "待定", "", ""),
    (84, "Round of 32 - 第12场", "2026-06-30", "待定", "待定", "", ""),
    (85, "Round of 32 - 第13场", "2026-07-01", "待定", "待定", "", ""),
    (86, "Round of 32 - 第14场", "2026-07-01", "待定", "待定", "", ""),
    (87, "Round of 32 - 第15场", "2026-07-01", "待定", "待定", "", ""),
    (88, "Round of 32 - 第16场", "2026-07-01", "待定", "待定", "", ""),
    (89, "Round of 16 - 第1场", "2026-07-02", "待定", "待定", "", ""),
    (90, "Round of 16 - 第2场", "2026-07-02", "待定", "待定", "", ""),
    (91, "Round of 16 - 第3场", "2026-07-02", "待定", "待定", "", ""),
    (92, "Round of 16 - 第4场", "2026-07-02", "待定", "待定", "", ""),
    (93, "Round of 16 - 第5场", "2026-07-03", "待定", "待定", "", ""),
    (94, "Round of 16 - 第6场", "2026-07-03", "待定", "待定", "", ""),
    (95, "Round of 16 - 第7场", "2026-07-03", "待定", "待定", "", ""),
    (96, "Round of 16 - 第8场", "2026-07-03", "待定", "待定", "", ""),
    (97, "Quarterfinal - 第1场", "2026-07-05", "待定", "待定", "", ""),
    (98, "Quarterfinal - 第2场", "2026-07-05", "待定", "待定", "", ""),
    (99, "Quarterfinal - 第3场", "2026-07-06", "待定", "待定", "", ""),
    (100, "Quarterfinal - 第4场", "2026-07-06", "待定", "待定", "", ""),
    (101, "Semifinal - 第1场", "2026-07-08", "待定", "待定", "", ""),
    (102, "Semifinal - 第2场", "2026-07-09", "待定", "待定", "", ""),
    (103, "3rd Place Match", "2026-07-11", "待定", "待定", "", ""),
    (104, "Final", "2026-07-12", "待定", "待定", "", "")
]

matches.extend(knockout)

# 写入比赛数据
for row_idx, match in enumerate(matches, 2):
    for col_idx, value in enumerate(match, 1):
        cell = ws_schedule.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws_schedule.row_dimensions[row_idx].height = 18

# 3. 参与者填写（优化：比赛作为行，参与者作为列）
ws_pred = wb.create_sheet("参与者填写")

# 标题行
ws_pred.cell(row=1, column=1, value="比赛编号").font = header_font
ws_pred.cell(row=1, column=1).fill = header_fill
ws_pred.cell(row=1, column=1).alignment = center_align
ws_pred.cell(row=1, column=1).border = thin_border

ws_pred.cell(row=1, column=2, value="队伍A").font = header_font
ws_pred.cell(row=1, column=2).fill = header_fill
ws_pred.cell(row=1, column=2).alignment = center_align
ws_pred.cell(row=1, column=2).border = thin_border

ws_pred.cell(row=1, column=3, value="队伍B").font = header_font
ws_pred.cell(row=1, column=3).fill = header_fill
ws_pred.cell(row=1, column=3).alignment = center_align
ws_pred.cell(row=1, column=3).border = thin_border

ws_pred.cell(row=1, column=4, value="实际比分A").font = header_font
ws_pred.cell(row=1, column=4).fill = header_fill
ws_pred.cell(row=1, column=4).alignment = center_align
ws_pred.cell(row=1, column=4).border = thin_border

ws_pred.cell(row=1, column=5, value="实际比分B").font = header_font
ws_pred.cell(row=1, column=5).fill = header_fill
ws_pred.cell(row=1, column=5).alignment = center_align
ws_pred.cell(row=1, column=5).border = thin_border

# 参与者列（最多10个参与者）
participants = ["张三", "李四", "王五", "赵六", "钱七", "孙八", "周九", "吴十"]
for p_idx, participant in enumerate(participants):
    col = p_idx + 6
    ws_pred.cell(row=1, column=col, value=f"{participant}预测A").font = header_font
    ws_pred.cell(row=1, column=col).fill = header_fill
    ws_pred.cell(row=1, column=col).alignment = center_align
    ws_pred.cell(row=1, column=col).border = thin_border
    
    ws_pred.cell(row=1, column=col+1, value=f"{participant}预测B").font = header_font
    ws_pred.cell(row=1, column=col+1).fill = header_fill
    ws_pred.cell(row=1, column=col+1).alignment = center_align
    ws_pred.cell(row=1, column=col+1).border = thin_border

ws_pred.row_dimensions[1].height = 30
ws_pred.column_dimensions['A'].width = 12
ws_pred.column_dimensions['B'].width = 18
ws_pred.column_dimensions['C'].width = 18
ws_pred.column_dimensions['D'].width = 12
ws_pred.column_dimensions['E'].width = 12

# 填充比赛行
for row_idx, match in enumerate(matches, 2):
    ws_pred.cell(row=row_idx, column=1, value=match[0]).alignment = center_align
    ws_pred.cell(row=row_idx, column=1).border = thin_border
    
    ws_pred.cell(row=row_idx, column=2, value=match[3]).alignment = center_align
    ws_pred.cell(row=row_idx, column=2).border = thin_border
    
    ws_pred.cell(row=row_idx, column=3, value=match[4]).alignment = center_align
    ws_pred.cell(row=row_idx, column=3).border = thin_border
    
    # 实际比分（从比赛赛程表自动获取）
    ws_pred.cell(row=row_idx, column=4).value = f"=比赛赛程!F{row_idx}"
    ws_pred.cell(row=row_idx, column=4).alignment = center_align
    ws_pred.cell(row=row_idx, column=4).border = thin_border
    
    ws_pred.cell(row=row_idx, column=5).value = f"=比赛赛程!G{row_idx}"
    ws_pred.cell(row=row_idx, column=5).alignment = center_align
    ws_pred.cell(row=row_idx, column=5).border = thin_border
    
    ws_pred.row_dimensions[row_idx].height = 20

# 4. 积分统计
ws_points = wb.create_sheet("积分统计")
ws_points.cell(row=1, column=1, value="参与者").font = header_font
ws_points.cell(row=1, column=1).fill = header_fill
ws_points.cell(row=1, column=1).alignment = center_align
ws_points.cell(row=1, column=1).border = thin_border
ws_points.column_dimensions['A'].width = 15
ws_points.row_dimensions[1].height = 30

# 创建104场比赛的积分列
for i in range(1, 105):
    col = i + 1
    ws_points.cell(row=1, column=col, value=f"比赛{i}").font = header_font
    ws_points.cell(row=1, column=col).fill = header_fill
    ws_points.cell(row=1, column=col).alignment = center_align
    ws_points.cell(row=1, column=col).border = thin_border
    ws_points.column_dimensions[get_column_letter(col)].width = 8

# 添加参与者和积分公式
for p_idx, participant in enumerate(participants):
    row = p_idx + 2
    ws_points.cell(row=row, column=1, value=participant).font = normal_font
    ws_points.cell(row=row, column=1).alignment = left_align
    ws_points.cell(row=row, column=1).border = thin_border
    ws_points.row_dimensions[row].height = 20
    
    # 为每场比赛添加积分计算公式
    for match in range(1, 105):
        col = match + 1
        pred_col_a = 6 + p_idx * 2
        pred_col_b = 7 + p_idx * 2
        
        # 积分公式
        formula = f'=IF(AND(参与者填写!{get_column_letter(pred_col_a)}{match+1}<>"",参与者填写!{get_column_letter(pred_col_b)}{match+1}<>""),IF(AND(参与者填写!{get_column_letter(pred_col_a)}{match+1}=参与者填写!D{match+1},参与者填写!{get_column_letter(pred_col_b)}{match+1}=参与者填写!E{match+1}),4,IF(SIGN(参与者填写!{get_column_letter(pred_col_a)}{match+1}-参与者填写!{get_column_letter(pred_col_b)}{match+1})=SIGN(参与者填写!D{match+1}-参与者填写!E{match+1}),1,0)),"")'
        
        ws_points.cell(row=row, column=col).value = formula
        ws_points.cell(row=row, column=col).alignment = center_align
        ws_points.cell(row=row, column=col).border = thin_border

# 5. 总排名
ws_ranking = wb.create_sheet("总排名")
headers = ["排名", "参与者", "总积分", "准确比分", "猜中胜负"]
ws_ranking.row_dimensions[1].height = 30

for col, header in enumerate(headers, 1):
    cell = ws_ranking.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws_ranking.column_dimensions['A'].width = 10
ws_ranking.column_dimensions['B'].width = 15
ws_ranking.column_dimensions['C'].width = 12
ws_ranking.column_dimensions['D'].width = 12
ws_ranking.column_dimensions['E'].width = 12

# 排名公式
for row_idx in range(2, 2 + len(participants)):
    ws_ranking.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
    ws_ranking.cell(row=row_idx, column=1).border = thin_border
    
    ws_ranking.cell(row=row_idx, column=2, value=f"=积分统计!A{row_idx}").alignment = center_align
    ws_ranking.cell(row=row_idx, column=2).border = thin_border
    
    ws_ranking.cell(row=row_idx, column=3, value=f"=SUM(积分统计!B{row_idx}:积分统计!CW{row_idx})").alignment = center_align
    ws_ranking.cell(row=row_idx, column=3).border = thin_border
    
    ws_ranking.cell(row=row_idx, column=4, value=f"=COUNTIF(积分统计!B{row_idx}:积分统计!CW{row_idx},4)").alignment = center_align
    ws_ranking.cell(row=row_idx, column=4).border = thin_border
    
    ws_ranking.cell(row=row_idx, column=5, value=f"=COUNTIF(积分统计!B{row_idx}:积分统计!CW{row_idx},1)").alignment = center_align
    ws_ranking.cell(row=row_idx, column=5).border = thin_border
    
    ws_ranking.row_dimensions[row_idx].height = 20

# 保存
wb.save("2026年世界杯竞猜-104场完整版.xlsx")
print(f"✅ 文件创建成功！包含 {len(matches)} 场比赛")
