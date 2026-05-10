#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "2026世界杯竞猜"

# 样式定义
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
title_font = Font(bold=True, size=13, color='366092')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 标题
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = "2026年足球世界杯竞猜活动"
title_cell.font = title_font
title_cell.alignment = center_align
ws.row_dimensions[1].height = 30

# 积分规则
ws.merge_cells('A2:E2')
rule_cell = ws['A2']
rule_cell.value = "积分规则：准确比分=4分 | 猜中胜负=1分 | 未猜中=0分"
rule_cell.font = Font(size=10, color='FF0000', bold=True)
rule_cell.alignment = center_align
ws.row_dimensions[2].height = 20

# 表头
headers = ["编号", "对阵", "日期", "实际比分", "你的预测"]
ws.row_dimensions[3].height = 35

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 设置列宽
col_widths = [8, 35, 15, 12, 12]
for idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = width

# 创建104场比赛数据
matches = []

# 小组赛：12组 x 6场 = 72场
groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for group in groups:
    for match_num in range(1, 7):
        match_id = len(matches) + 1
        # 计算日期（从6月11日开始）
        base_date = 10 + (match_id - 1) // 8
        date_str = f"2026-06-{base_date:02d}"
        matches.append([match_id, f"小组赛{group}组", date_str])

# 淘汰赛：32场
knockout_matches = [
    (73, "1/32决赛-1", "2026-06-28"),
    (74, "1/32决赛-2", "2026-06-28"),
    (75, "1/32决赛-3", "2026-06-28"),
    (76, "1/32决赛-4", "2026-06-28"),
    (77, "1/32决赛-5", "2026-06-29"),
    (78, "1/32决赛-6", "2026-06-29"),
    (79, "1/32决赛-7", "2026-06-29"),
    (80, "1/32决赛-8", "2026-06-29"),
    (81, "1/32决赛-9", "2026-06-30"),
    (82, "1/32决赛-10", "2026-06-30"),
    (83, "1/32决赛-11", "2026-06-30"),
    (84, "1/32决赛-12", "2026-06-30"),
    (85, "1/32决赛-13", "2026-07-01"),
    (86, "1/32决赛-14", "2026-07-01"),
    (87, "1/32决赛-15", "2026-07-01"),
    (88, "1/32决赛-16", "2026-07-01"),
    (89, "1/16决赛-1", "2026-07-02"),
    (90, "1/16决赛-2", "2026-07-02"),
    (91, "1/16决赛-3", "2026-07-03"),
    (92, "1/16决赛-4", "2026-07-03"),
    (93, "1/16决赛-5", "2026-07-04"),
    (94, "1/16决赛-6", "2026-07-04"),
    (95, "1/16决赛-7", "2026-07-05"),
    (96, "1/16决赛-8", "2026-07-05"),
    (97, "1/4决赛-1", "2026-07-07"),
    (98, "1/4决赛-2", "2026-07-07"),
    (99, "1/4决赛-3", "2026-07-08"),
    (100, "1/4决赛-4", "2026-07-08"),
    (101, "半决赛-1", "2026-07-10"),
    (102, "半决赛-2", "2026-07-11"),
    (103, "季军赛", "2026-07-13"),
    (104, "决赛", "2026-07-14"),
]

for ko in knockout_matches:
    matches.append([ko[0], ko[1], ko[2]])

print(f"总比赛场数：{len(matches)}")

# 填充表格数据
for idx, match in enumerate(matches, 4):
    row = idx
    
    # 编号
    cell = ws.cell(row=row, column=1, value=match[0])
    cell.alignment = center_align
    cell.border = thin_border
    
    # 对阵（简化显示）
    cell = ws.cell(row=row, column=2, value=match[1])
    cell.alignment = center_align
    cell.border = thin_border
    
    # 日期
    cell = ws.cell(row=row, column=3, value=match[2])
    cell.alignment = center_align
    cell.border = thin_border
    
    # 实际比分（黄色标记，管理员填写）
    cell = ws.cell(row=row, column=4)
    cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border
    
    # 预测（用户填写）
    cell = ws.cell(row=row, column=5)
    cell.alignment = center_align
    cell.border = thin_border
    
    ws.row_dimensions[row].height = 18

# 总分统计
total_row = 4 + len(matches)
ws.cell(row=total_row, column=1, value="你的总积分").font = Font(bold=True, size=11)
ws.merge_cells(f'A{total_row}:D{total_row}')
cell = ws.cell(row=total_row, column=1)
cell.fill = header_fill
cell.alignment = center_align

# 总分单元格
total_cell = ws.cell(row=total_row, column=5)
total_cell.value = f'=SUM(E4:E{total_row-1})'
total_cell.font = Font(bold=True, size=14, color='FF0000')
total_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
total_cell.alignment = center_align
total_cell.border = thin_border
ws.row_dimensions[total_row].height = 30

# 使用说明
note_row = total_row + 2
ws.cell(row=note_row, column=1, value="【使用说明】").font = Font(bold=True, size=10)
ws.merge_cells(f'A{note_row}:E{note_row}')

notes = [
    "1. 黄色单元格由管理员填写实际比分（格式：2-1）",
    "2. 在'预测'列填写你的预测（格式：2-1）",
    "3. 积分规则：准确比分=4分，猜中胜负=1分，未猜中=0分",
    "4. 底部会自动计算你的总积分",
    "5. 建议每人复制一份文件，填写完成后提交给管理员统计"
]

for i, note in enumerate(notes):
    row = note_row + 1 + i
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=note).font = Font(size=9)

print(f"✅ Excel文件创建成功！共 {len(matches)} 场比赛")
wb.save("2026世界杯竞猜-单表版.xlsx")
print("文件路径：/workspace/2026世界杯竞猜-单表版.xlsx")
