from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "2026世界杯竞猜"

# 样式定义
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
title_font = Font(bold=True, size=14, color='366092')
normal_font = Font(size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws.merge_cells('A1:L1')
ws['A1'] = "2026年足球世界杯竞猜活动 - 预测与积分统计表"
ws['A1'].font = title_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 30

# 积分规则说明
ws['A2'] = "积分规则：准确比分→4分 | 只猜中胜负→1分 | 未猜中→0分"
ws['A2'].font = Font(size=10, color='FF0000', bold=True)
ws.merge_cells('A2:L2')
ws.row_dimensions[2].height = 20

# 表头（第3行）
headers = ["比赛编号", "比赛阶段", "日期", "队伍A", "队伍B", "实际比分A", "实际比分B", "张三预测A", "张三预测B", "张三积分", "李四预测A", "李四预测B", "李四积分"]
ws.row_dimensions[3].height = 40

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 设置列宽
col_widths = [10, 20, 12, 15, 15, 10, 10, 10, 10, 8, 10, 10, 8]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 创建104场比赛数据
matches = []

# 小组赛：12组×6场 = 72场
groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for group in groups:
    for match in range(1, 7):
        match_id = len(matches) + 1
        matches.append([
            match_id,
            f"小组赛{group}组第{match}轮",
            f"2026-06-{10 + (match_id-1)//8:02d}",
            f"{group}队1",
            f"{group}队2",
            None,  # 实际比分A
            None   # 实际比分B
        ])

# 淘汰赛：32场
knockout = [
    (73, "Round of 32-1", "2026-06-28", "待定", "待定", None, None),
    (74, "Round of 32-2", "2026-06-28", "待定", "待定", None, None),
    (75, "Round of 32-3", "2026-06-28", "待定", "待定", None, None),
    (76, "Round of 32-4", "2026-06-28", "待定", "待定", None, None),
    (77, "Round of 32-5", "2026-06-29", "待定", "待定", None, None),
    (78, "Round of 32-6", "2026-06-29", "待定", "待定", None, None),
    (79, "Round of 32-7", "2026-06-29", "待定", "待定", None, None),
    (80, "Round of 32-8", "2026-06-29", "待定", "待定", None, None),
    (81, "Round of 32-9", "2026-06-30", "待定", "待定", None, None),
    (82, "Round of 32-10", "2026-06-30", "待定", "待定", None, None),
    (83, "Round of 32-11", "2026-06-30", "待定", "待定", None, None),
    (84, "Round of 32-12", "2026-06-30", "待定", "待定", None, None),
    (85, "Round of 32-13", "2026-07-01", "待定", "待定", None, None),
    (86, "Round of 32-14", "2026-07-01", "待定", "待定", None, None),
    (87, "Round of 32-15", "2026-07-01", "待定", "待定", None, None),
    (88, "Round of 32-16", "2026-07-01", "待定", "待定", None, None),
    (89, "Round of 16-1", "2026-07-02", "待定", "待定", None, None),
    (90, "Round of 16-2", "2026-07-02", "待定", "待定", None, None),
    (91, "Round of 16-3", "2026-07-02", "待定", "待定", None, None),
    (92, "Round of 16-4", "2026-07-02", "待定", "待定", None, None),
    (93, "Round of 16-5", "2026-07-03", "待定", "待定", None, None),
    (94, "Round of 16-6", "2026-07-03", "待定", "待定", None, None),
    (95, "Round of 16-7", "2026-07-03", "待定", "待定", None, None),
    (96, "Round of 16-8", "2026-07-03", "待定", "待定", None, None),
    (97, "Quarterfinal-1", "2026-07-05", "待定", "待定", None, None),
    (98, "Quarterfinal-2", "2026-07-05", "待定", "待定", None, None),
    (99, "Quarterfinal-3", "2026-07-06", "待定", "待定", None, None),
    (100, "Quarterfinal-4", "2026-07-06", "待定", "待定", None, None),
    (101, "Semifinal-1", "2026-07-08", "待定", "待定", None, None),
    (102, "Semifinal-2", "2026-07-09", "待定", "待定", None, None),
    (103, "3rd Place", "2026-07-11", "待定", "待定", None, None),
    (104, "Final", "2026-07-12", "待定", "待定", None, None),
]

matches.extend(knockout)

# 参与者列表
participants = ["张三", "李四"]

# 填充比赛数据
for row_idx, match in enumerate(matches, 4):
    # 比赛编号
    ws.cell(row=row_idx, column=1, value=match[0]).alignment = center_align
    ws.cell(row=row_idx, column=1).border = thin_border
    
    # 比赛阶段
    ws.cell(row=row_idx, column=2, value=match[1]).alignment = center_align
    ws.cell(row=row_idx, column=2).border = thin_border
    
    # 日期
    ws.cell(row=row_idx, column=3, value=match[2]).alignment = center_align
    ws.cell(row=row_idx, column=3).border = thin_border
    
    # 队伍A
    ws.cell(row=row_idx, column=4, value=match[3]).alignment = center_align
    ws.cell(row=row_idx, column=4).border = thin_border
    
    # 队伍B
    ws.cell(row=row_idx, column=5, value=match[4]).alignment = center_align
    ws.cell(row=row_idx, column=5).border = thin_border
    
    # 实际比分A
    cell_f = ws.cell(row=row_idx, column=6)
    cell_f.alignment = center_align
    cell_f.border = thin_border
    cell_f.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    
    # 实际比分B
    cell_g = ws.cell(row=row_idx, column=7)
    cell_g.alignment = center_align
    cell_g.border = thin_border
    cell_g.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    
    # 参与者预测和积分
    for p_idx, participant in enumerate(participants):
        base_col = 8 + p_idx * 3
        
        # 预测比分A
        pred_a = ws.cell(row=row_idx, column=base_col)
        pred_a.alignment = center_align
        pred_a.border = thin_border
        
        # 预测比分B
        pred_b = ws.cell(row=row_idx, column=base_col+1)
        pred_b.alignment = center_align
        pred_b.border = thin_border
        
        # 积分计算公式
        # 4分：准确比分
        # 1分：胜负关系正确
        # 0分：都不对
       积分_formula = f'=IF(AND(F{row_idx}<>"",G{row_idx}<>""),IF(AND({get_column_letter(base_col)}{row_idx}=F{row_idx},{get_column_letter(base_col+1)}{row_idx}=G{row_idx}),4,IF(SIGN({get_column_letter(base_col)}{row_idx}-{get_column_letter(base_col+1)}{row_idx})=SIGN(F{row_idx}-G{row_idx}),1,0)),"")'
        
       积分_cell = ws.cell(row=row_idx, column=base_col+2)
        积分_cell.value = 积分_formula
        积分_cell.alignment = center_align
        积分_cell.border = thin_border
        积分_cell.font = Font(bold=True, size=10)
    
    ws.row_dimensions[row_idx].height = 18

# 添加总积分行
total_row = 4 + len(matches)
ws.cell(row=total_row, column=1, value="总分").font = Font(bold=True, size=11)
ws.cell(row=total_row, column=1).fill = header_fill
ws.cell(row=total_row, column=1).font = header_font
ws.cell(row=total_row, column=1).alignment = center_align
ws.cell(row=total_row, column=1).border = thin_border

# 合并总分行的其他列
for col in range(2, 8):
    ws.cell(row=total_row, column=col).border = thin_border

# 为每个参与者添加总分公式
for p_idx, participant in enumerate(participants):
    base_col = 8 + p_idx * 3
    
    # 参与者姓名
    ws.cell(row=total_row, column=base_col, value=participant).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=base_col).alignment = center_align
    ws.cell(row=total_row, column=base_col).border = thin_border
    
    ws.cell(row=total_row, column=base_col+1).border = thin_border
    
    # 总分
    total_formula = f'=SUM({get_column_letter(base_col+2)}4:{get_column_letter(base_col+2)}{total_row-1})'
    total_cell = ws.cell(row=total_row, column=base_col+2)
    total_cell.value = total_formula
    total_cell.font = Font(bold=True, size=12, color='FF0000')
    total_cell.alignment = center_align
    total_cell.border = thin_border
    total_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

ws.row_dimensions[total_row].height = 25

# 添加使用说明
note_row = total_row + 2
ws.cell(row=note_row, column=1, value="使用说明：").font = Font(bold=True, size=11)
ws.merge_cells(f'A{note_row}:L{note_row}')

note_row += 1
notes = [
    "1. 黄色单元格（F、G列）由管理员填写实际比赛比分",
    "2. 参与者在对应列填写预测比分（如：2表示2球）",
    "3. 积分会自动计算：准确比分→4分，只猜中胜负→1分，未猜中→0分",
    "4. 表格底部会自动统计每位参与者的总积分",
    "5. 可将文件分享到微信群，使用在线Excel可实现多人同时填写"
]
for i, note in enumerate(notes):
    ws.cell(row=note_row+i, column=1, value=note).font = Font(size=10)
    ws.merge_cells(f'A{note_row+i}:L{note_row+i}')

print(f"✅ 单工作表版本创建成功！包含 {len(matches)} 场比赛")
print(f"   参与者：{', '.join(participants)}")
wb.save("2026年世界杯竞猜-单表版.xlsx")
