from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# 定义样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
title_font = Font(bold=True, size=14, color='366092')
normal_font = Font(size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. 使用说明工作表
ws_guide = wb.active
ws_guide.title = "使用说明"
ws_guide.column_dimensions['A'].width = 60
ws_guide.column_dimensions['B'].width = 40

guide_content = [
    ("2026年足球世界杯竞猜活动 - 使用说明", "title"),
    ("", "blank"),
    ("📋 工作表说明", "header"),
    ("1. 【比赛赛程】- 包含所有比赛信息，管理员填写实际比赛结果", "normal"),
    ("2. 【参与者填写】- 参与者在此填写每场比赛的预测", "normal"),
    ("3. 【积分统计】- 自动计算每位参与者的积分", "normal"),
    ("4. 【排名】- 自动生成参与者积分排名", "normal"),
    ("", "blank"),
    ("🎯 积分规则", "header"),
    ("• 预测胜负平正确：得 3 分", "normal"),
    ("• 预测准确比分：得 5 分（包含胜负平和比分的双重奖励）", "normal"),
    ("• 预测胜负平错误：得 0 分", "normal"),
    ("", "blank"),
    ("📝 使用步骤", "header"),
    ("1. 在【比赛赛程】表中填写所有比赛信息", "normal"),
    ("2. 将Excel文件分享到微信群", "normal"),
    ("3. 每位参与者在【参与者填写】表中填写自己的预测", "normal"),
    ("4. 比赛结束后，在【比赛赛程】表中填写实际结果", "normal"),
    ("5. 【积分统计】和【排名】表会自动更新", "normal"),
    ("", "blank"),
    ("⚠️ 注意事项", "header"),
    ("• 每位参与者使用不同的行填写预测", "normal"),
    ("• 预测必须在比赛开始前填写", "normal"),
    ("• 实际结果由管理员统一填写", "normal"),
    ("• 文件可同时供多人编辑（需使用在线Excel）", "normal"),
]

row = 1
for text, style in guide_content:
    cell = ws_guide.cell(row=row, column=1, value=text)
    if style == "title":
        cell.font = Font(bold=True, size=16, color='366092')
        ws_guide.row_dimensions[row].height = 30
    elif style == "header":
        cell.font = Font(bold=True, size=12, color='366092')
        ws_guide.row_dimensions[row].height = 25
    elif style == "normal":
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws_guide.row_dimensions[row].height = 20
    elif style == "blank":
        ws_guide.row_dimensions[row].height = 10
    row += 1

# 2. 比赛赛程工作表
ws_schedule = wb.create_sheet("比赛赛程")
headers = ["比赛编号", "比赛阶段", "比赛日期", "比赛时间", "队伍A", "队伍B", "实际结果-队伍A", "实际结果-队伍B", "实际胜负平"]
ws_schedule.row_dimensions[1].height = 30

for col, header in enumerate(headers, 1):
    cell = ws_schedule.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 设置列宽
ws_schedule.column_dimensions['A'].width = 12
ws_schedule.column_dimensions['B'].width = 15
ws_schedule.column_dimensions['C'].width = 12
ws_schedule.column_dimensions['D'].width = 10
ws_schedule.column_dimensions['E'].width = 20
ws_schedule.column_dimensions['F'].width = 20
ws_schedule.column_dimensions['G'].width = 15
ws_schedule.column_dimensions['H'].width = 15
ws_schedule.column_dimensions['I'].width = 12

# 添加示例比赛数据（小组赛示例）
sample_matches = [
    (1, "小组赛A组第1轮", "2026-06-11", "20:00", "墨西哥", "波兰", "", "", ""),
    (2, "小组赛A组第1轮", "2026-06-11", "23:00", "阿根廷", "沙特", "", "", ""),
    (3, "小组赛B组第1轮", "2026-06-12", "20:00", "美国", "威尔士", "", "", ""),
    (4, "小组赛B组第1轮", "2026-06-12", "23:00", "英格兰", "伊朗", "", "", ""),
    (5, "小组赛C组第1轮", "2026-06-13", "20:00", "阿根廷", "墨西哥", "", "", ""),
    (6, "小组赛C组第1轮", "2026-06-13", "23:00", "波兰", "沙特", "", "", ""),
]

for row_idx, match in enumerate(sample_matches, 2):
    for col_idx, value in enumerate(match, 1):
        cell = ws_schedule.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws_schedule.row_dimensions[row_idx].height = 20

# 添加数据验证（实际胜负平）
dv_result = DataValidation(type="list", formula1='"胜,平,负"', allow_blank=True)
ws_schedule.add_data_validation(dv_result)
dv_result.add("I2:I105")

# 3. 参与者填写工作表
ws_prediction = wb.create_sheet("参与者填写")
headers = ["参与者姓名", "比赛1预测-队伍A", "比赛1预测-队伍B", "比赛1预测-胜负平", 
           "比赛2预测-队伍A", "比赛2预测-队伍B", "比赛2预测-胜负平",
           "比赛3预测-队伍A", "比赛3预测-队伍B", "比赛3预测-胜负平",
           "比赛4预测-队伍A", "比赛4预测-队伍B", "比赛4预测-胜负平",
           "比赛5预测-队伍A", "比赛5预测-队伍B", "比赛5预测-胜负平",
           "比赛6预测-队伍A", "比赛6预测-队伍B", "比赛6预测-胜负平"]

ws_prediction.row_dimensions[1].height = 30
for col, header in enumerate(headers, 1):
    cell = ws_prediction.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 设置列宽
ws_prediction.column_dimensions['A'].width = 15
for col in range(2, 20):
    ws_prediction.column_dimensions[get_column_letter(col)].width = 12

# 添加数据验证（胜负平预测）
dv_pred = DataValidation(type="list", formula1='"胜,平,负"', allow_blank=True)
ws_prediction.add_data_validation(dv_pred)
for i in range(1, 7):
    start_col = (i-1)*3 + 4
    dv_pred.add(f"{get_column_letter(start_col)}2:{get_column_letter(start_col)}100")

# 添加示例参与者行
sample_participants = [
    ("张三", 2, 1, "胜", 1, 0, "胜", 2, 1, "胜", 3, 1, "胜", 2, 0, "胜", 1, 0, "胜"),
    ("李四", 1, 1, "平", 2, 1, "胜", 1, 2, "负", 2, 2, "平", 1, 1, "平", 2, 1, "胜"),
    ("王五", 0, 2, "负", 1, 2, "负", 3, 0, "胜", 1, 3, "负", 0, 2, "负", 1, 1, "平"),
]

for row_idx, participant in enumerate(sample_participants, 2):
    for col_idx, value in enumerate(participant, 1):
        cell = ws_prediction.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws_prediction.row_dimensions[row_idx].height = 20

# 4. 积分统计工作表
ws_points = wb.create_sheet("积分统计")
headers = ["参与者姓名", "总积分", "比赛1积分", "比赛2积分", "比赛3积分", 
           "比赛4积分", "比赛5积分", "比赛6积分"]

ws_points.row_dimensions[1].height = 30
for col, header in enumerate(headers, 1):
    cell = ws_points.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws_points.column_dimensions['A'].width = 15
for col in range(2, 9):
    ws_points.column_dimensions[get_column_letter(col)].width = 12

# 添加公式（示例）
for row_idx in range(2, 5):
    # 总积分公式
    ws_points.cell(row=row_idx, column=2).value = f"=SUM(C{row_idx}:H{row_idx})"
    
    # 各场比赛积分公式（需要根据实际逻辑计算）
    for match in range(1, 7):
        col_idx = match + 2
        # 这里需要复杂的逻辑来判断预测是否正确
        # 简化版：检查预测胜负平是否与实际一致
        pred_col = (match-1)*3 + 4
        score_a_col = (match-1)*3 + 2
        score_b_col = (match-1)*3 + 3
        
        formula = f'=IF(参与者填写!{get_column_letter(pred_col)}{row_idx}="","",IF(参与者填写!{get_column_letter(pred_col)}{row_idx}=比赛赛程!I{match+1},3,0)+IF(AND(参与者填写!{get_column_letter(score_a_col)}{row_idx}=比赛赛程!G{match+1},参与者填写!{get_column_letter(score_b_col)}{row_idx}=比赛赛程!H{match+1}),2,0))'
        ws_points.cell(row=row_idx, column=col_idx).value = formula

# 设置样式
for row_idx in range(2, 5):
    for col_idx in range(1, 9):
        cell = ws_points.cell(row=row_idx, column=col_idx)
        if col_idx == 1:
            cell.alignment = left_align
        else:
            cell.alignment = center_align
        cell.font = normal_font
        cell.border = thin_border
    ws_points.row_dimensions[row_idx].height = 20

# 5. 排名工作表
ws_ranking = wb.create_sheet("排名")
headers = ["排名", "参与者姓名", "总积分"]

ws_ranking.row_dimensions[1].height = 30
for col, header in enumerate(headers, 1):
    cell = ws_ranking.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws_ranking.column_dimensions['A'].width = 10
ws_ranking.column_dimensions['B'].width = 20
ws_ranking.column_dimensions['C'].width = 15

# 添加排名公式
for row_idx in range(2, 5):
    ws_ranking.cell(row=row_idx, column=1).value = row_idx - 1
    ws_ranking.cell(row=row_idx, column=2).value = f"=积分统计!A{row_idx}"
    ws_ranking.cell(row=row_idx, column=3).value = f"=积分统计!B{row_idx}"

# 设置样式
for row_idx in range(2, 5):
    for col_idx in range(1, 4):
        cell = ws_ranking.cell(row=row_idx, column=col_idx)
        cell.alignment = center_align
        cell.font = normal_font
        cell.border = thin_border
    ws_ranking.row_dimensions[row_idx].height = 20

# 保存文件
wb.save("2026年世界杯竞猜活动.xlsx")
print("Excel文件创建成功：2026年世界杯竞猜活动.xlsx")
