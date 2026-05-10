from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# 定义样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
title_font = Font(bold=True, size=14, color='366092')
normal_font = Font(size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. 使用说明工作表
ws_guide = wb.active
ws_guide.title = "使用说明"
ws_guide.column_dimensions['A'].width = 80

guide_content = [
    "2026年足球世界杯竞猜活动 - 使用说明",
    "",
    "📋 工作表说明",
    "1. 【比赛赛程】- 包含所有104场比赛信息，管理员填写实际比赛结果",
    "2. 【参与者填写】- 参与者在此填写每场比赛的比分预测",
    "3. 【积分统计】- 自动计算每位参与者每场比赛的积分",
    "4. 【总排名】- 自动生成参与者总积分排名",
    "",
    "🎯 积分规则（新）",
    "• 预测准确比分：得 4 分",
    "• 只预测对胜负平关系：得 1 分",
    "• 胜负关系和比分都没猜中：得 0 分",
    "",
    "📝 使用步骤",
    "1. 在【比赛赛程】表中填写所有104场比赛的信息",
    "2. 将Excel文件分享到微信群（建议使用在线Excel）",
    "3. 每位参与者在【参与者填写】表中填写自己的预测",
    "4. 比赛结束后，在【比赛赛程】表中填写实际比分",
    "5. 【积分统计】和【总排名】表会自动更新",
    "",
    "⚠️ 注意事项",
    "• 每位参与者使用不同的行填写预测",
    "• 预测必须在比赛开始前填写",
    "• 实际比分由管理员统一填写",
    "• 文件可同时供多人编辑（需使用在线Excel或腾讯文档）",
    "• 本表格已预设104场比赛，涵盖全部赛程"
]

for row, text in enumerate(guide_content, 1):
    cell = ws_guide.cell(row=row, column=1, value=text)
    if row == 1:
        cell.font = Font(bold=True, size=16, color='366092')
        ws_guide.row_dimensions[row].height = 30
    elif text.startswith("📋") or text.startswith("🎯") or text.startswith("📝") or text.startswith("⚠️"):
        cell.font = Font(bold=True, size=12, color='366092')
        ws_guide.row_dimensions[row].height = 25
    else:
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws_guide.row_dimensions[row].height = 20

# 2. 比赛赛程工作表（104场比赛）
ws_schedule = wb.create_sheet("比赛赛程")
headers = ["比赛编号", "比赛阶段", "比赛日期", "队伍A", "队伍B", "实际比分A", "实际比分B", "实际胜负平"]
ws_schedule.row_dimensions[1].height = 30

for col, header in enumerate(headers, 1):
    cell = ws_schedule.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 设置列宽
ws_schedule.column_dimensions['A'].width = 12
ws_schedule.column_dimensions['B'].width = 20
ws_schedule.column_dimensions['C'].width = 15
ws_schedule.column_dimensions['D'].width = 20
ws_schedule.column_dimensions['E'].width = 20
ws_schedule.column_dimensions['F'].width = 12
ws_schedule.column_dimensions['G'].width = 12
ws_schedule.column_dimensions['H'].width = 12

# 创建104场比赛的占位符
# 2026年世界杯：48队，104场比赛
# 小组赛：72场（12组×6场）
# 淘汰赛：32场（16+8+4+2+1+1）

matches = []

# 小组赛A-L组，每组6场比赛
groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for group_idx, group in enumerate(groups):
    for match in range(1, 7):
        match_id = len(matches) + 1
        matches.append([
            match_id,
            f"小组赛{group}组第{match}轮",
            f"2026-06-{11 + (match_id-1)//8:02d}",
            f"待定{group}队1",
            f"待定{group}队2",
            "",
            "",
            ""
        ])

# 淘汰赛阶段
knockout_matches = [
    (73, "Round of 32 - 第1场", "2026-06-28", "待定", "待定", "", "", ""),
    (74, "Round of 32 - 第2场", "2026-06-28", "待定", "待定", "", "", ""),
    (75, "Round of 32 - 第3场", "2026-06-29", "待定", "待定", "", "", ""),
    (76, "Round of 32 - 第4场", "2026-06-29", "待定", "待定", "", "", ""),
    (77, "Round of 32 - 第5场", "2026-06-30", "待定", "待定", "", "", ""),
    (78, "Round of 32 - 第6场", "2026-06-30", "待定", "待定", "", "", ""),
    (79, "Round of 32 - 第7场", "2026-07-01", "待定", "待定", "", "", ""),
    (80, "Round of 32 - 第8场", "2026-07-01", "待定", "待定", "", "", ""),
    (81, "Round of 16 - 第1场", "2026-07-02", "待定", "待定", "", "", ""),
    (82, "Round of 16 - 第2场", "2026-07-02", "待定", "待定", "", "", ""),
    (83, "Round of 16 - 第3场", "2026-07-03", "待定", "待定", "", "", ""),
    (84, "Round of 16 - 第4场", "2026-07-03", "待定", "待定", "", "", ""),
    (85, "Round of 16 - 第5场", "2026-07-04", "待定", "待定", "", "", ""),
    (86, "Round of 16 - 第6场", "2026-07-04", "待定", "待定", "", "", ""),
    (87, "Round of 16 - 第7场", "2026-07-05", "待定", "待定", "", "", ""),
    (88, "Round of 16 - 第8场", "2026-07-05", "待定", "待定", "", "", ""),
    (89, "Quarterfinal - 第1场", "2026-07-07", "待定", "待定", "", "", ""),
    (90, "Quarterfinal - 第2场", "2026-07-07", "待定", "待定", "", "", ""),
    (91, "Quarterfinal - 第3场", "2026-07-08", "待定", "待定", "", "", ""),
    (92, "Quarterfinal - 第4场", "2026-07-08", "待定", "待定", "", "", ""),
    (93, "Semifinal - 第1场", "2026-07-10", "待定", "待定", "", "", ""),
    (94, "Semifinal - 第2场", "2026-07-11", "待定", "待定", "", "", ""),
    (95, "3rd Place Match", "2026-07-13", "待定", "待定", "", "", ""),
    (96, "Final", "2026-07-14", "待定", "待定", "", "", ""),
]

matches.extend(knockout_matches)

# 填充比赛数据
for row_idx, match_data in enumerate(matches, 2):
    for col_idx, value in enumerate(match_data, 1):
        cell = ws_schedule.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    ws_schedule.row_dimensions[row_idx].height = 18

# 3. 参与者填写工作表（优化布局）
ws_pred = wb.create_sheet("参与者填写")
ws_pred.column_dimensions['A'].width = 15
ws_pred.column_dimensions['B'].width = 12
ws_pred.column_dimensions['C'].width = 12

# 标题行
ws_pred.cell(row=1, column=1, value="参与者姓名").font = header_font
ws_pred.cell(row=1, column=1).fill = header_fill
ws_pred.cell(row=1, column=1).alignment = center_align
ws_pred.cell(row=1, column=1).border = thin_border
ws_pred.row_dimensions[1].height = 30

# 为前20场比赛创建列标题
for i in range(1, 21):
    col_offset = (i-1) * 3 + 1
    ws_pred.cell(row=1, column=col_offset+1, value=f"比赛{i}-队伍A比分").font = header_font
    ws_pred.cell(row=1, column=col_offset+1).fill = header_fill
    ws_pred.cell(row=1, column=col_offset+1).alignment = center_align
    ws_pred.cell(row=1, column=col_offset+1).border = thin_border
    
    ws_pred.cell(row=1, column=col_offset+2, value=f"比赛{i}-队伍B比分").font = header_font
    ws_pred.cell(row=1, column=col_offset+2).fill = header_fill
    ws_pred.cell(row=1, column=col_offset+2).alignment = center_align
    ws_pred.cell(row=1, column=col_offset+2).border = thin_border
    
    ws_pred.column_dimensions[get_column_letter(col_offset+1)].width = 12
    ws_pred.column_dimensions[get_column_letter(col_offset+2)].width = 12

# 添加示例参与者
participants = ["张三", "李四", "王五", "赵六"]
for p_idx, participant in enumerate(participants, 2):
    ws_pred.cell(row=p_idx, column=1, value=participant).font = normal_font
    ws_pred.cell(row=p_idx, column=1).alignment = left_align
    ws_pred.cell(row=p_idx, column=1).border = thin_border
    ws_pred.row_dimensions[p_idx].height = 20
    
    # 填充示例预测数据（前20场）
    import random
    for i in range(1, 21):
        col_offset = (i-1) * 3 + 1
        score_a = random.randint(0, 4)
        score_b = random.randint(0, 4)
        ws_pred.cell(row=p_idx, column=col_offset+1, value=score_a).alignment = center_align
        ws_pred.cell(row=p_idx, column=col_offset+1).border = thin_border
        ws_pred.cell(row=p_idx, column=col_offset+2, value=score_b).alignment = center_align
        ws_pred.cell(row=p_idx, column=col_offset+2).border = thin_border

# 4. 积分统计工作表
ws_points = wb.create_sheet("积分统计")
ws_points.cell(row=1, column=1, value="参与者姓名").font = header_font
ws_points.cell(row=1, column=1).fill = header_fill
ws_points.cell(row=1, column=1).alignment = center_align
ws_points.cell(row=1, column=1).border = thin_border
ws_points.column_dimensions['A'].width = 15
ws_points.row_dimensions[1].height = 30

# 创建104场比赛的积分列
for i in range(1, 105):
    col = i + 1
    ws_points.cell(row=1, column=col, value=f"比赛{i}积分").font = header_font
    ws_points.cell(row=1, column=col).fill = header_fill
    ws_points.cell(row=1, column=col).alignment = center_align
    ws_points.cell(row=1, column=col).border = thin_border
    ws_points.column_dimensions[get_column_letter(col)].width = 10

# 添加参与者和积分计算公式
for p_idx, participant in enumerate(participants, 2):
    ws_points.cell(row=p_idx, column=1, value=participant).font = normal_font
    ws_points.cell(row=p_idx, column=1).alignment = left_align
    ws_points.cell(row=p_idx, column=1).border = thin_border
    ws_points.row_dimensions[p_idx].height = 20
    
    # 为每场比赛添加积分计算公式
    for match in range(1, 105):
        col = match + 1
        
        # 获取预测比分和实际比分的列
        pred_col_a = get_column_letter((match-1) * 3 + 2)
        pred_col_b = get_column_letter((match-1) * 3 + 3)
        
        # 实际比分在比赛赛程表的F和G列
        actual_col_a = f"比赛赛程!F{match+1}"
        actual_col_b = f"比赛赛程!G{match+1}"
        
        # 积分公式：
        # 4分：准确比分
        # 1分：胜负平关系正确
        # 0分：都不对
        formula = f'=IF(AND(参与者填写!{pred_col_a}{p_idx}<>"",参与者填写!{pred_col_b}{p_idx}<>""),IF(AND(参与者填写!{pred_col_a}{p_idx}={actual_col_a},参与者填写!{pred_col_b}{p_idx}={actual_col_b}),4,IF(SIGN(参与者填写!{pred_col_a}{p_idx}-参与者填写!{pred_col_b}{p_idx})=SIGN({actual_col_a}-{actual_col_b}),1,0)),"")'
        
        ws_points.cell(row=p_idx, column=col).value = formula
        ws_points.cell(row=p_idx, column=col).alignment = center_align
        ws_points.cell(row=p_idx, column=col).border = thin_border

# 5. 总排名工作表
ws_ranking = wb.create_sheet("总排名")
headers = ["排名", "参与者姓名", "总积分", "准确比分场数", "猜中胜负平场数", "总预测场数"]
ws_ranking.row_dimensions[1].height = 30

for col, header in enumerate(headers, 1):
    cell = ws_ranking.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws_ranking.column_dimensions['A'].width = 10
ws_ranking.column_dimensions['B'].width = 15
ws_ranking.column_dimensions['C'].width = 12
ws_ranking.column_dimensions['D'].width = 15
ws_ranking.column_dimensions['E'].width = 15
ws_ranking.column_dimensions['F'].width = 15

# 添加排名公式
for row_idx in range(2, 2 + len(participants)):
    ws_ranking.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align
    ws_ranking.cell(row=row_idx, column=1).border = thin_border
    
    ws_ranking.cell(row=row_idx, column=2, value=f"=积分统计!A{row_idx}").alignment = center_align
    ws_ranking.cell(row=row_idx, column=2).border = thin_border
    
    # 总积分 = SUM of all match points
    ws_ranking.cell(row=row_idx, column=3, value=f"=SUM(积分统计!B{row_idx}:积分统计!CW{row_idx})").alignment = center_align
    ws_ranking.cell(row=row_idx, column=3).border = thin_border
    
    # 准确比分场数（4分）
    ws_ranking.cell(row=row_idx, column=4, value=f"=COUNTIF(积分统计!B{row_idx}:积分统计!CW{row_idx},4)").alignment = center_align
    ws_ranking.cell(row=row_idx, column=4).border = thin_border
    
    # 猜中胜负平场数（1分）
    ws_ranking.cell(row=row_idx, column=5, value=f"=COUNTIF(积分统计!B{row_idx}:积分统计!CW{row_idx},1)").alignment = center_align
    ws_ranking.cell(row=row_idx, column=5).border = thin_border
    
    # 总预测场数
    ws_ranking.cell(row=row_idx, column=6, value=f"=COUNT(积分统计!B{row_idx}:积分统计!CW{row_idx})").alignment = center_align
    ws_ranking.cell(row=row_idx, column=6).border = thin_border
    
    ws_ranking.row_dimensions[row_idx].height = 20

# 保存文件
wb.save("2026年世界杯竞猜活动-完整版.xlsx")
print("完整版Excel文件创建成功：2026年世界杯竞猜活动-完整版.xlsx")
print(f"包含 {len(matches)} 场比赛")
