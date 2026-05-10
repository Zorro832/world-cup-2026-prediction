#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "2026世界杯竞猜"

# 样式
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
title_font = Font(bold=True, size=13, color='366092')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 标题
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = "2026年足球世界杯竞猜 - 填写你的预测"
title_cell.font = title_font
title_cell.alignment = center_align
ws.row_dimensions[1].height = 30

# 积分规则
ws.merge_cells('A2:E2')
rule_cell = ws['A2']
rule_cell.value = "积分规则：准确比分4分 | 猜中胜负1分 | 未猜中0分"
rule_cell.font = Font(size=10, color='FF0000', bold=True)
rule_cell.alignment = center_align
ws.row_dimensions[2].height = 20

# 表头
headers = ["比赛编号", "对阵双方", "比赛日期", "实际比分", "你的预测"]
ws.row_dimensions[3].height = 35

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 列宽
col_widths = [10, 30, 15, 12, 12]
for idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = width

# 创建104场比赛（简化版）
matches = []

# 小组赛 72场
groups = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
for group in groups:
    for m in range(1, 7):
        mid = len(matches) + 1
        date = f"2026-06-{10+(mid-1)//8:02d}"
        matches.append([mid, f"{group}组", date, "队伍A", "队伍B"])

# 淘汰赛 32场
knockout_stages = [
    (73, "Round of 32-1", "2026-06-28"),
    (74, "Round of 32-2", "2026-06-28"),
    (81, "R16-1", "2026-07-02"),
    (85, "QF-1", "2026-07-05"),
    (89, "SF-1", "2026-07-08"),
    (91, "3rd Place", "2026-07-11"),
    (92, "Final", "2026-07-12"),
]

# 补齐104场
current_id = 73
for ko in knockout_stages:
    while current_id <= ko[0] and current_id <= 104:
        matches.append([current_id, ko[1], ko[2], "TBD", "TBD"])
        current_id += 1

# 填充表格
for idx, match in enumerate(matches, 4):
    row = idx
    
    # 编号
    cell = ws.cell(row=row, column=1, value=match[0])
    cell.alignment = center_align
    cell.border = thin_border
    
    # 对阵
    cell = ws.cell(row=row, column=2, value=f"{match[3]} vs {match[4]}")
    cell.alignment = center_align
    cell.border = thin_border
    
    # 日期
    cell = ws.cell(row=row, column=3, value=match[2])
    cell.alignment = center_align
    cell.border = thin_border
    
    # 实际比分（黄色，管理员填写）
    for col in [4]:
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    
    # 预测（用户填写）
    for col in [5]:
        cell = ws.cell(row=row, column=col)
        cell.alignment = center_align
        cell.border = thin_border
    
    ws.row_dimensions[row].height = 18

# 总分
total_row = 4 + len(matches)
ws.cell(row=total_row, column=1, value="你的最终得分").font = Font(bold=True, size=11)
ws.merge_cells(f'A{total_row}:D{total_row}')
cell = ws.cell(row=total_row, column=1)
cell.fill = header_fill
cell.alignment = center_align

# 总分公式（简化：直接求和预测列）
total_cell = ws.cell(row=total_row, column=5)
total_cell.value = f'=SUM(E4:E{total_row-1})'
total_cell.font = Font(bold=True, size=14, color='FF0000')
total_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
total_cell.alignment = center_align
total_cell.border = thin_border
ws.row_dimensions[total_row].height = 30

# 使用说明
note_row = total_row + 2
ws.cell(row=note_row, column=1, value="使用说明：").font = Font(bold=True, size=10)
ws.merge_cells(f'A{note_row}:E{note_row}')

notes = [
    "1. 黄色单元格由管理员在赛后填写实际比分",
    "2. 在'你的预测'列填写预测比分（格式：2-1）",
    "3. 积分需手动计算或请管理员帮忙设置公式",
    "4. 每人复制一份文件填写，完成后提交给管理员"
]
for i, note in enumerate(notes):
    row = note_row + 1 + i
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=note).font = Font(size=9)

print(f"✅ 创建成功！共 {len(matches)} 场比赛")
wb.save("2026世界杯竞猜-单表版.xlsx")
print("文件已保存：2026世界杯竞猜-单表版.xlsx")
